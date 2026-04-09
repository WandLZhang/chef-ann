"""
@file load_district_users.py
@brief Parse xlsx files to extract authorized user emails and build district mapping.

@details Reads three xlsx files from Downloads:
1. TX UXR Districts List.xlsx - 108 selected districts with CE IDs
2. UXR Study TX Districts.xlsx - Master District List with emails
3. WBSCM Orders Feb 2026.xlsx - Entitlement/order data

Outputs:
- Authorized user emails (CN Directors + Superintendents)
- District lookup (email → CE ID, district, county, enrollment)
- District entitlement summary (district → entitlement, spent, remaining)
- Cross-reference check between all three files

@author Willis Zhang
@date 2026-04-09
"""

import json
import os
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

DOWNLOADS = os.path.expanduser("~/Downloads")


def load_districts_list():
    """Parse TX UXR Districts List.xlsx - the 108 selected districts with CE IDs."""
    path = os.path.join(DOWNLOADS, "TX UXR Districts List.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Sheet1"]

    districts = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header row
        if not row[0] or str(row[0]).strip().lower() == "total":
            continue

        account_name = str(row[0]).strip()
        ce_id = row[1] if len(row) > 1 else None
        city = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        state = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        zip_code = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        county = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        num_schools = row[6] if len(row) > 6 else None
        enrollment = row[7] if len(row) > 7 else None
        email = str(row[8]).strip() if len(row) > 8 and row[8] else ""

        districts[account_name] = {
            "account_name": account_name,
            "ce_id": str(ce_id) if ce_id else "",
            "city": city,
            "state": state,
            "zip": zip_code,
            "county": county,
            "num_schools": num_schools,
            "enrollment": enrollment,
            "associated_email": email,
        }

    wb.close()
    print("\n=== TX UXR Districts List ===")
    print(f"Total districts: {len(districts)}")
    print(
        f"Districts with emails: {sum(1 for d in districts.values() if d['associated_email'])}"
    )
    return districts


def load_master_district_list():
    """Parse UXR Study TX Districts.xlsx - Master District List with CN Director/Superintendent emails."""
    path = os.path.join(DOWNLOADS, "UXR Study TX Districts.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)

    # Parse Selected Districts sheet first
    ws_selected = wb["Selected Districts"]
    selected_names = set()
    for i, row in enumerate(ws_selected.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[0] and str(row[0]).strip().lower() != "total":
            selected_names.add(str(row[0]).strip())

    # Parse Master District List sheet
    ws_master = wb["Master District List"]
    headers = None

    # Build email-to-district mapping
    email_to_district = {}
    district_emails = defaultdict(
        lambda: {
            "cn_emails": set(),
            "super_emails": set(),
            "ce_name": "",
            "sites": [],
            "grades": set(),
        }
    )
    all_ce_names = set()

    for i, row in enumerate(ws_master.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
            print(f"\nMaster District List headers: {headers}")
            continue

        ce_name = str(row[0]).strip() if row[0] else ""
        if not ce_name:
            continue

        all_ce_names.add(ce_name)

        # Find email columns by header index
        super_email_idx = (
            headers.index("SuperintendentEmail")
            if "SuperintendentEmail" in headers
            else None
        )
        cn_email_idx = (
            headers.index("ChildNutDirEmail") if "ChildNutDirEmail" in headers else None
        )
        site_name_idx = headers.index("SiteName") if "SiteName" in headers else None
        county_idx = headers.index("CECounty") if "CECounty" in headers else None

        super_email = (
            str(row[super_email_idx]).strip().lower()
            if super_email_idx and row[super_email_idx]
            else ""
        )
        cn_email = (
            str(row[cn_email_idx]).strip().lower()
            if cn_email_idx and row[cn_email_idx]
            else ""
        )
        site_name = (
            str(row[site_name_idx]).strip()
            if site_name_idx and row[site_name_idx]
            else ""
        )
        county = str(row[county_idx]).strip() if county_idx and row[county_idx] else ""

        district_emails[ce_name]["ce_name"] = ce_name
        if cn_email and cn_email != "none":
            district_emails[ce_name]["cn_emails"].add(cn_email)
            email_to_district[cn_email] = ce_name
        if super_email and super_email != "none":
            district_emails[ce_name]["super_emails"].add(super_email)
            email_to_district[super_email] = ce_name
        if site_name:
            district_emails[ce_name]["sites"].append(site_name)
        if county:
            district_emails[ce_name]["county"] = county

        # Extract grade levels from Grade* columns (Y/N per school site)
        grade_col_names = [h for h in headers if h and h.startswith("Grade")]
        for gcol in grade_col_names:
            gidx = headers.index(gcol)
            gval = row[gidx] if gidx < len(row) else None
            if gval and str(gval).upper() in ("Y", "TRUE", "YES", "1"):
                district_emails[ce_name]["grades"].add(gcol)

    wb.close()

    # Collect all unique emails
    all_emails = set()
    for d in district_emails.values():
        all_emails.update(d["cn_emails"])
        all_emails.update(d["super_emails"])

    print("\n=== Master District List ===")
    print(f"Total unique CE Names (ISDs): {len(all_ce_names)}")
    print(f"CE Names in Selected Districts: {len(selected_names)}")
    print(
        f"Unique CN Director emails: {sum(len(d['cn_emails']) for d in district_emails.values())}"
    )
    print(
        f"Unique Superintendent emails: {sum(len(d['super_emails']) for d in district_emails.values())}"
    )
    print(f"Total unique emails: {len(all_emails)}")

    return district_emails, email_to_district, all_ce_names, selected_names


def load_wbscm_orders():
    """Parse WBSCM Orders Feb 2026.xlsx - entitlement data per district."""
    path = os.path.join(DOWNLOADS, "WBSCM Orders Feb 2026.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["PY27-Entitlements"]

    district_entitlements = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header row
        if not row[0]:
            continue

        sold_to_name = str(row[2]).strip() if row[2] else ""
        sold_to_party = str(row[1]).strip() if row[1] else ""

        if not sold_to_name or sold_to_name in district_entitlements:
            # Already captured this district's summary - just track total spend
            if sold_to_name in district_entitlements:
                net_value = float(row[19]) if row[19] else 0
                district_entitlements[sold_to_name]["total_ordered"] += net_value
                district_entitlements[sold_to_name]["order_count"] += 1
            continue

        # First occurrence - capture entitlement info
        # [28] = Beginning Balance, [29] = DOD Fresh Amount
        beginning_balance = float(row[28]) if len(row) > 28 and row[28] else 0
        dod_fresh = float(row[29]) if len(row) > 29 and row[29] else 0
        net_value = float(row[19]) if len(row) > 19 and row[19] else 0

        district_entitlements[sold_to_name] = {
            "sold_to_name": sold_to_name,
            "sold_to_party": sold_to_party,
            "beginning_balance": beginning_balance,
            "dod_fresh_amount": dod_fresh,
            "total_ordered": net_value,
            "order_count": 1,
        }

    # Now re-scan to get accurate total_ordered for each district
    # Reset and re-calculate
    district_totals = defaultdict(lambda: {"total_ordered": 0, "order_count": 0})
    ws2 = wb["PY27-Entitlements"]
    for i, row in enumerate(ws2.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row[0]:
            continue
        sold_to_name = str(row[2]).strip() if row[2] else ""
        net_value = float(row[19]) if row[19] else 0
        district_totals[sold_to_name]["total_ordered"] += net_value
        district_totals[sold_to_name]["order_count"] += 1

    # Merge totals
    for name, totals in district_totals.items():
        if name in district_entitlements:
            district_entitlements[name]["total_ordered"] = round(
                totals["total_ordered"], 2
            )
            district_entitlements[name]["order_count"] = totals["order_count"]

    wb.close()

    print("\n=== WBSCM Orders PY27 ===")
    print(f"Total districts with orders: {len(district_entitlements)}")
    print(f"Sample districts: {list(district_entitlements.keys())[:5]}")

    return district_entitlements


def cross_reference(districts_list, master_emails, master_ce_names, wbscm_orders):
    """Cross-reference all three data sources and check for AGUA DULCE specifically."""

    print("\n=== CROSS-REFERENCE CHECK ===")

    # Check AGUA DULCE specifically
    print("\n--- AGUA DULCE CHECK ---")
    print(
        f"In TX UXR Districts List (108 selected): {'AGUA DULCE' in str(districts_list.keys())}"
    )

    # Check master district list
    agua_dulce_in_master = [
        name for name in master_ce_names if "AGUA DULCE" in name.upper()
    ]
    print(f"In Master District List: {agua_dulce_in_master}")

    # Check WBSCM
    agua_dulce_in_wbscm = [
        name for name in wbscm_orders.keys() if "AGUA DULCE" in name.upper()
    ]
    print(f"In WBSCM Orders: {agua_dulce_in_wbscm}")
    if agua_dulce_in_wbscm:
        for name in agua_dulce_in_wbscm:
            ent = wbscm_orders[name]
            print(
                f"  Entitlement: ${ent['beginning_balance']:,.2f}, DOD Fresh: ${ent['dod_fresh_amount']:,.2f}, Total Ordered: ${ent['total_ordered']:,.2f}"
            )

    # Check which selected districts are in WBSCM
    selected_in_wbscm = 0
    selected_not_in_wbscm = []
    wbscm_names_upper = {name.upper(): name for name in wbscm_orders.keys()}

    for dist_name in districts_list.keys():
        # Try exact match and fuzzy
        name_upper = dist_name.upper()
        # Remove "Independent School District" → try matching with "ISD"
        short_name = (
            name_upper.replace("INDEPENDENT SCHOOL DISTRICT", "ISD")
            .replace("  ", " ")
            .strip()
        )

        found = False
        for wbscm_upper, wbscm_orig in wbscm_names_upper.items():
            if name_upper in wbscm_upper or wbscm_upper in name_upper:
                found = True
                break
            if short_name in wbscm_upper or wbscm_upper in short_name:
                found = True
                break

        if found:
            selected_in_wbscm += 1
        else:
            selected_not_in_wbscm.append(dist_name)

    print("\n--- Selected Districts in WBSCM ---")
    print(f"Matched: {selected_in_wbscm} / {len(districts_list)}")
    if selected_not_in_wbscm:
        print(
            f"Not matched ({len(selected_not_in_wbscm)}): {selected_not_in_wbscm[:10]}..."
        )

    # Check which WBSCM districts are NOT in the selected 108
    wbscm_not_selected = []
    districts_upper = {name.upper() for name in districts_list.keys()}
    for wbscm_name in wbscm_orders.keys():
        name_upper = wbscm_name.upper()
        # Try to match by expanding ISD → Independent School District
        expanded = name_upper.replace(" ISD", " INDEPENDENT SCHOOL DISTRICT")
        if name_upper not in districts_upper and expanded not in districts_upper:
            wbscm_not_selected.append(wbscm_name)

    print(f"\nWBSCM districts NOT in selected 108: {len(wbscm_not_selected)}")
    if wbscm_not_selected:
        print(f"  Examples: {wbscm_not_selected[:10]}")


def build_district_lookup(districts_list, master_emails, wbscm_orders):
    """Build the email → district → entitlement lookup."""

    # Build name matching between districts_list and WBSCM
    wbscm_name_map = {}
    for wbscm_name in wbscm_orders.keys():
        wbscm_name_map[wbscm_name.upper()] = wbscm_name

    # Build name matching between districts_list account names and master CE names
    # districts_list uses "Austin Independent School District"
    # master uses "AUSTIN ISD" (sometimes)
    # WBSCM uses "AUSTIN ISD" (short form)

    lookup = {}

    for ce_name, data in master_emails.items():
        # All emails for this ISD
        all_emails = data["cn_emails"] | data["super_emails"]

        # Try to find CE ID from districts_list
        ce_id = ""
        district_info = None

        # Try matching CE name to districts_list Account Name
        for acct_name, dist in districts_list.items():
            # Match strategies:
            # 1. Master: "BROWNSVILLE ISD" ↔ Districts: "Brownsville Independent School District"
            acct_upper = acct_name.upper()
            ce_upper = ce_name.upper()

            # Normalize: expand ISD, strip extra spaces around dashes
            ce_expanded = ce_upper.replace(" ISD", " INDEPENDENT SCHOOL DISTRICT")
            acct_short = (
                acct_upper.replace("INDEPENDENT SCHOOL DISTRICT", "ISD")
                .replace("  ", " ")
                .strip()
            )
            # Also normalize dashes: " - " → "-" for matching
            ce_norm = ce_upper.replace(" - ", "-").replace("- ", "-").replace(" -", "-")
            acct_norm = (
                acct_short.replace(" - ", "-").replace("- ", "-").replace(" -", "-")
            )
            ce_exp_norm = (
                ce_expanded.replace(" - ", "-").replace("- ", "-").replace(" -", "-")
            )
            acct_upper_norm = (
                acct_upper.replace(" - ", "-").replace("- ", "-").replace(" -", "-")
            )

            if (
                ce_upper == acct_upper
                or ce_expanded == acct_upper
                or ce_upper == acct_short
                or ce_norm == acct_norm
                or ce_exp_norm == acct_upper_norm
            ):
                ce_id = dist["ce_id"]
                district_info = dist
                break

        # Try to find WBSCM entitlement
        entitlement = None
        ce_upper = ce_name.upper()
        # WBSCM uses short names like "AGUA DULCE ISD"
        for wbscm_upper, wbscm_orig in wbscm_name_map.items():
            if ce_upper in wbscm_upper or wbscm_upper in ce_upper:
                entitlement = wbscm_orders[wbscm_orig]
                break
            # Also try ISD expansion
            ce_as_isd = ce_upper
            if "ISD" not in ce_upper:
                ce_as_isd = ce_upper + " ISD"
            if ce_as_isd in wbscm_upper or wbscm_upper == ce_as_isd:
                entitlement = wbscm_orders[wbscm_orig]
                break

        # Map raw grade columns to app grade bands
        raw_grades = data.get("grades", set())
        grade_bands = []
        if any(
            g in raw_grades
            for g in ["GradeEarlyEducation", "GradeHeadStart", "GradePreK"]
        ):
            grade_bands.append("prek")
        if any(
            g in raw_grades
            for g in ["GradeKinder", "Grade1", "Grade2", "Grade3", "Grade4", "Grade5"]
        ):
            grade_bands.append("elementary")
        if any(g in raw_grades for g in ["Grade6", "Grade7", "Grade8"]):
            grade_bands.append("middle")
        if any(g in raw_grades for g in ["Grade9", "Grade10", "Grade11", "Grade12"]):
            grade_bands.append("high")

        for email in all_emails:
            lookup[email] = {
                "email": email,
                "ce_name": ce_name,
                "ce_id": ce_id,
                "county": data.get(
                    "county", district_info["county"] if district_info else ""
                ),
                "num_schools": district_info["num_schools"] if district_info else None,
                "enrollment": district_info["enrollment"] if district_info else None,
                "role": "cn_director"
                if email in data["cn_emails"]
                else "superintendent",
                "entitlement": entitlement,
                "grade_bands": grade_bands,
            }

    return lookup


def main():
    print("=" * 60)
    print("LOADING DISTRICT DATA FROM XLSX FILES")
    print("=" * 60)

    # 1. Load all three files
    districts_list = load_districts_list()
    master_emails, email_to_district, all_ce_names, selected_names = (
        load_master_district_list()
    )
    wbscm_orders = load_wbscm_orders()

    # 2. Cross-reference (including AGUA DULCE check)
    cross_reference(districts_list, master_emails, all_ce_names, wbscm_orders)

    # 3. Build lookup
    lookup = build_district_lookup(districts_list, master_emails, wbscm_orders)

    # 4. Collect all unique emails for APPROVED_USERS
    all_emails = sorted(set(lookup.keys()))

    print("\n=== FINAL OUTPUT ===")
    print(f"Total unique authorized emails: {len(all_emails)}")

    # 5. Save outputs
    output_dir = os.path.join(os.path.dirname(__file__), "..", "functions", "data")

    # Save email list
    email_list_path = os.path.join(output_dir, "authorized_emails.json")
    with open(email_list_path, "w") as f:
        json.dump(all_emails, f, indent=2)
    print(f"Saved {len(all_emails)} emails to {email_list_path}")

    # Save district lookup (email → district info)
    lookup_path = os.path.join(output_dir, "district_lookup.json")
    # Convert for JSON serialization
    lookup_serializable = {}
    for email, info in lookup.items():
        entry = dict(info)
        # Clean float→int for numeric fields from openpyxl
        if entry.get("num_schools") is not None:
            entry["num_schools"] = int(entry["num_schools"])
        if entry.get("enrollment") is not None:
            entry["enrollment"] = int(entry["enrollment"])
        ce_id_raw = str(entry.get("ce_id", ""))
        entry["ce_id"] = (
            ce_id_raw.replace(".0", "") if ce_id_raw.endswith(".0") else ce_id_raw
        )
        # Remove entitlement sub-object or flatten it
        if entry.get("entitlement"):
            entry["entitlement_amount"] = entry["entitlement"]["beginning_balance"]
            entry["dod_fresh_amount"] = entry["entitlement"]["dod_fresh_amount"]
            entry["total_ordered"] = entry["entitlement"]["total_ordered"]
            entry["sold_to_party"] = entry["entitlement"]["sold_to_party"]
        del entry["entitlement"]
        lookup_serializable[email] = entry

    with open(lookup_path, "w") as f:
        json.dump(lookup_serializable, f, indent=2)
    print(f"Saved district lookup to {lookup_path}")

    # Save WBSCM entitlement summary
    entitlement_path = os.path.join(output_dir, "district_entitlements.json")
    with open(entitlement_path, "w") as f:
        json.dump(wbscm_orders, f, indent=2, default=str)
    print(f"Saved {len(wbscm_orders)} district entitlements to {entitlement_path}")

    # Print sample for verification
    print("\n--- Sample Authorized Emails (first 20) ---")
    for email in all_emails[:20]:
        info = lookup[email]
        ent_str = (
            f"${info.get('entitlement', {}).get('beginning_balance', 0):,.0f}"
            if info.get("entitlement")
            else "N/A"
        )
        print(
            f"  {email:<45} → {info['ce_name']:<35} CE:{info['ce_id']:<6} Ent:{ent_str}"
        )

    # Print the APPROVED_USERS array for AuthContext.tsx
    print("\n--- APPROVED_USERS for AuthContext.tsx ---")
    # Include existing users + new ones
    existing = [
        "williszhang@google.com",
        "ngoren@google.com",
        "lori.nelson@chefannfoundation.org",
        "test@williszhang.altostrat.com",
    ]
    combined = sorted(set(existing + all_emails))
    print("const APPROVED_USERS = [")
    for email in combined:
        print(f"  '{email}',")
    print("];")
    print(f"Total: {len(combined)} users")


if __name__ == "__main__":
    main()
